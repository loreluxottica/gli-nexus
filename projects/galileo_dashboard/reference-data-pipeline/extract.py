"""Extract sheets of Galileo Frontend.xlsx into JSON.

Prefers archived workbooks under archive/claude-work/ and falls back to the
canonical source workbook in data/source/.
"""
import json
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]


def _newest_excel():
    candidates = list((ROOT / "archive" / "claude-work").rglob("Galileo Frontend.xlsx"))
    candidates += list((ROOT / "data" / "source").glob("Galileo Frontend.xlsx"))
    if not candidates:
        raise FileNotFoundError("Galileo Frontend.xlsx not found")
    return max(candidates, key=lambda p: p.stat().st_mtime)


SRC = _newest_excel()
OUT_DIR = ROOT / "data"
OUT_DIR.mkdir(exist_ok=True)


def sheet_to_rows(ws):
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([("" if v is None else str(v)).strip() for v in row])
    while rows and all(c == "" for c in rows[-1]):
        rows.pop()
    return rows


def main():
    print(f"Source: {SRC}")
    wb = load_workbook(SRC, data_only=True, read_only=True)
    print("Sheets:", wb.sheetnames)

    out = {"source_file": SRC.name, "sheets": []}
    for idx, name in enumerate(wb.sheetnames):
        ws = wb[name]
        rows = sheet_to_rows(ws)
        out["sheets"].append({
            "index": idx,
            "name": name,
            "row_count": len(rows),
            "col_count": max((len(r) for r in rows), default=0),
            "rows": rows,
        })

    (OUT_DIR / "raw.json").write_text(json.dumps(out, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"Wrote {OUT_DIR / 'raw.json'}")

    for sheet in out["sheets"]:
        print(f"\n=== Sheet {sheet['index']}: {sheet['name']} ({sheet['row_count']} x {sheet['col_count']}) ===")
        for row in sheet["rows"][:30]:
            print(row)


if __name__ == "__main__":
    main()
